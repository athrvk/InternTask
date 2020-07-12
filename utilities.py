# Previous attempt
# Useless now I switched to xlwings

import requests
import datetime
import openpyxl as px


class intern_test:
    api_key = "7cac41f62c2d8231a68bbf5697ffea5a"

    flag = 1

    def __init__(self, city_name='Indore', unit_system='metric'):
        self.city_name = city_name
        self.unit_system = unit_system
        self.url = "http://api.openweathermap.org/data/2.5/weather?q={}&units={}&appid={}".format(self.city_name,
                                                                                                  self.unit_system,
                                                                                                  self.api_key)

    def req_data(self):
        requested_data = requests.get(self.url)
        print("Data request complete...")
        return requested_data.json()

    def create_xlfile(self, filename='test.xlsx'):
        workbook = px.Workbook()
        workbook.save(filename=filename)
        print("Excel file : {} created...".format(filename))
        return workbook

    def load_xlfile(self, filename='test.xlsx'):
        wb = px.load_workbook(filename)
        print("Workbook loaded...")
        return wb

    def initialize_workbook(self, workbook):
        sh1 = workbook.create_sheet("Sheet1")
        sh2 = workbook.create_sheet("Sheet2")
        std = workbook.get_sheet_by_name("Sheet")
        workbook.remove_sheet(std)
        columns = ['Date/Time', "City Name", "Temp", "Option C/F", 'Option End']
        sh1.append(columns)
        sh2.append(["City Name"])
        print("Workbook initialized...")

    def append_data(self, data, workbook):
        dt = datetime.datetime.now()
        c_name = data['name']
        temp = data['main']['temp']
        sheet1 = workbook['Sheet1']
        sheet1.append([dt, c_name, temp])
        print("Data appended successfully...")

    def save_file(self, workbook, name):
        workbook.save(filename=name)
        print("Workbook saved as {}...".format(name))

    def check_temp(self, workbook):
        sh1 = workbook['Sheet1']
        option = ''
        if sh1 is not None:
            if sh1['D2'].value is not None:
                option = str(sh1["D2"].value.lower())
        if option == 'c':
            self.unit_system = 'metric'
        if option == 'f':
            self.unit_system = 'imperial'

    def check_state(self, workbook):
        sh1 = workbook['Sheet1']
        option = ''
        if sh1 is not None:
            if sh1['F2'].value is not None:
                option = int(sh1['F2'].value)
        if option == 1:
            self.flag = 1
        if option == 0:
            self.flag = 0
